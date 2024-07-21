import openpyxl as op
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import urllib
import pyautogui
import random
from selenium.webdriver import Chrome

driver = Chrome()
driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
  "source": """
    Object.defineProperty(navigator, 'webdriver', {
      get: () => undefined
    })
  """
})
driver.get("https://passport.weibo.com/sso/signin?entry=miniblog&source=miniblog")  # 打开微博登录界面
time.sleep(10)
print("success")
""" url = "https://s.weibo.com/weibo?q=%E9%95%BF%E6%B2%99%E6%9A%B4%E9%9B%A8&scope=ori&suball=1&timescope=custom%3A2024-06-25%3A2024-06-26&Refer=g"
driver.get(url)
 """

def getPageNumMax():
    pageNumMax = 0  #搜索结果页数
    try:
        pageNum = driver.find_element(By.CLASS_NAME,"s-scroll")
        pages = pageNum.find_elements(By.XPATH,"./li")
        for num in pages:
            pageNumMax += 1
        return(pageNumMax)
    except:return 1

def unfold():
    unfold = driver.find_elements(By.XPATH,"//*[text()='展开']")#展开所有页面内容
    for p in unfold:
        p.click()

def Existed(url,column):
    if url != None:
        for cell in column:
            if url == cell.value :
                print("已存在！")
                return True
        return False    
        
def clearRow(sheet,rowNum):
    for i in range(1,5):
        sheet.cell(rowNum,i, value = '')

def showBigImg():#展开图像
    classes = ('m1','m2','m3','m4')
    for i in classes:
        BigImg(i)

def BigImg(str):
    btn = driver.find_elements(By.XPATH,"//*[@class='card-feed']/div[2]/div[3]/div/ul/li[1]")
    for i in btn:
        try:
            i.click()
        except:print("unclickable")

def clickDownLoad(str):
    time.sleep(2)
    pyautogui.click(500,500,button = 'right')
    time.sleep(1)
    pyautogui.typewrite(['v'])
    time.sleep(1)
    pyautogui.typewrite(str)
    pyautogui.typewrite(['enter'])
    pyautogui.click(518,27,button='left')

def ImgCrawler(url,sheet,p):
    i=0
    row_num=sheet.max_row
    
    while i < getPageNumMax(): #搜索结果页数
        url = url+"&page="+str(i)
        driver.get(url)
        unfold()
        #showBigImg()

        Xpath = "//*[@class='main-full']/div/div[@class='card-wrap']"
        card = driver.find_elements(By.XPATH,Xpath)#所有推文卡
        print(driver.find_element(By.XPATH,Xpath).text)
        try:
            driver.find_element(By.XPATH,"//*[@class='main-full']/div[@class='card-wrap']")#存在该元素表明没有搜索结果
            i+=1
            continue
        
        except:

            for j in card:
                print(j.text)
                
                nickname = j.find_element(By.CLASS_NAME,"info") #用户名
                sheet.cell(row_num,1,value=nickname.text.replace('c\n',''))
                sheet.cell(row_num,5,value=str(p))
                row_num +=1
                
                tweet_time = j.find_element(By.CLASS_NAME,"from") #发布时间
                sheet.cell(row_num,2,value=tweet_time.text)
                
                txt = j.find_element(By.CLASS_NAME,"txt")# 发布内容 #推文内容
                sheet.cell(row_num,3,value=txt.text)             
                #time.sleep(5)
                #print(j.text)
                try:
                    j.find_element(By.XPATH,".//*[@class='content']/div[3]/div/ul/li").click() #查看第一张大图
                    pics = j.find_elements(By.XPATH,".//*[@class='choose-pic']/ul/li")
                    print(pics[0].text)
                    
                    
                    try:
                        for l in pics:
                            l.click()
                            time.sleep(random.randint(1,3))
                            j.find_element(By.XPATH,".//*[@class='tab']/li[2]/a").click()#查看大图
                            clickDownLoad(str(p))
                            p += 1
                        
                        time.sleep(random.randint(1,3))
                        j.find_element(By.XPATH,"//*[@class='tab']/li/a").click() #收起
                        time.sleep(random.randint(1,3))
                    except:
                        try:
                            j.find_element(By.XPATH,".//*[@class='tab']/li[2]/a").click()#查看大图
                            clickDownLoad(str(p))
                            p +=5
                        except:print("next!")
                except:print("Fight!") 

            """ pics = j.find_elements(By.CLASS_NAME,'cur')
            for p in pics:
                p.click
                pic = j.find_element(By.XPATH,".//*[@class='bigpic next']/div/img")
                picUrl=pic.get_attribute("src")
                print(picUrl)
                print(row_num)
                if Existed(picUrl,sheet['D']):#判断图片url是否已存在
                    clearRow(sheet,row_num)
                    break
                else:
                    sheet.cell(row_num,4,value=picUrl)
                    row_num +=1
            row_num +=1
            """
            


            """ pics = j.find_elements(By.XPATH,".//div/div/div/div/ul/li/img")#图片链接
            for n in pics:
                picUrl=n.get_attribute("src")
                print(picUrl)
                print(row_num)
                if Existed(picUrl,sheet['D']):#判断图片url是否已存在
                    clearRow(sheet,row_num)
                    break
                else:
                    sheet.cell(row_num,4,value=picUrl)
                    row_num +=1
            row_num +=1 """
            i = i + 1
    return p

def VidCrawler(url,sheet):
   
    i=0
    row_num = sheet.max_row
    while i < getPageNumMax(): #搜索结果页数
        url = url+"&page="+str(i)
        driver.get(url)
        unfold()
        #showBigImg()

        Xpath = "//*[@class='main-full']/div/div[@class='card-wrap']"
        card = driver.find_elements(By.XPATH,Xpath)#所有推文卡
        try :
            driver.find_element(By.XPATH,"//*[@class='main-full']/div[@class='card-wrap']")#存在该元素表明没有搜索结果，跳转至下一页
            i+=1 
            continue
        except:
            for j in card:
                    print(j.text)
                    
                    nickname = j.find_element(By.CLASS_NAME,"info") #用户名
                    sheet2.cell(row_num,1,value=nickname.text.replace('c\n',''))                
                
                    tweet_time = j.find_element(By.CLASS_NAME,"from") #发布时间
                    sheet2.cell(row_num,2,value=tweet_time.text)
                    
                    txt = j.find_element(By.CLASS_NAME,"txt")# 发布内容 #推文内容
                    sheet2.cell(row_num,3,value=txt.text)
                
                    videos = j.find_elements(By.XPATH,".//video")#视频链接
                    for u in videos:
                        sheet2.cell(row_num,4,value=u.get_attribute("src"))
                    row_num += 1
            i = i + 1

KeyWords = ('长沙暴雨','长沙洪水')
#book = op.load_workbook("C:/Users/ZH/Downloads/DL/Pic1.xlsx")  # 创建表格
#sheet = book.active
book2 = op.load_workbook("C:/Users/ZH/Downloads/DL/Vid1.xlsx")  # 创建表格
sheet2 = book2.active
p=1
for KW in KeyWords:
    for D in range(23,27):#日
        for J in range(23):
            #Imgurl = "https://s.weibo.com/weibo?q="+urllib.parse.quote(KW,'utf-8')+"&scope=ori&haspic=1&timescope=custom%3A2024-06-{}-{}%3A2024-06-{}-{}".format(D,J,D,J+1)
            Vidurl = "https://s.weibo.com/weibo?q="+urllib.parse.quote(KW,'utf-8')+"&scope=ori&haspic=0&timescope=custom%3A2024-06-{}-{}%3A2024-06-{}-{}".format(D,J,D,J+1)
            #p = ImgCrawler(Imgurl,sheet,p)
            VidCrawler(Vidurl,sheet2)
            #book.save("./Pic2.xlsx")
            book2.save("./Vid2.xlsx")     