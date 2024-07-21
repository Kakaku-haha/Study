import re
import openpyxl
import os
book = openpyxl.load_workbook("./IP2.xlsx")
sheet =book.active

findIP = re.compile(r'IP属地：(.*?)$')
findTime = re.compile(r'(.*?)来自')

for row in range(1,len(sheet['A'])):
    
    value = sheet.cell(row,7).value
    if value:
        #print(value)
        IP = re.findall(findIP,value)
        #print(IP)
        Time = re.findall(findTime,sheet.cell(row,2).value)
        print(Time)
        if not IP:
            continue
        if IP[0] != "湖南":
            delete_Name = sheet.cell(row,1).value
            print(delete_Name)
            try:
                os.remove("C:/Users/ZH/Downloads/vide/"+delete_Name+".mp4")
            except:print("Can't Find")
#book.save("./test.xlsx")