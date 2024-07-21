import selenium
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import urllib
import time
driver = webdriver.Chrome()
#url =""
#driver.get(url)
driver.get("https://passport.weibo.com/sso/signin?entry=miniblog&source=miniblog")  # 打开微博登录界面
time.sleep(10)
book = openpyxl.load_workbook("C:/Users/ZH/Downloads/DL/href1.xlsx")
sheet = book.active


for row in range(1,len(sheet['A'])):
    try:
        #if row % 300 == 0:
            #time.sleep(200)
        if sheet.cell(row,6).value != None:
            print("continue!")
            continue
        url ="https://s.weibo.com/user?q="+urllib.parse.quote(sheet.cell(row,1).value,'utf-8')+"&Refer=weibo_user#"
        
        driver.get(url)
        #time.sleep(2)
        user = driver.find_element(By.XPATH,"//*[@class='avator']/a")
        href = user.get_attribute("href")
        
        sheet.cell(row,6).value = href
        print(href)  
    except:
        print("Noooooh!")
        sheet.cell(row,6).value = None
    print(row)
    book.save("./href.xlsx")    