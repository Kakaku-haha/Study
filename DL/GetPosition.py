import re
import openpyxl

url = "C:/Users/ZH/Downloads/DL/Vid2.xlsx"
findPosition = re.compile(r'2(.*?)·(.*?) ')
#findPosition = re.compile(r'2(.*)·(.*)')
table = openpyxl.load_workbook(url)
ws = table.active


for c in range(1,len(ws['A'])):
    if(ws.cell(c,3).value !=None):
        #print(ws.cell(c,3).value)
        str = re.findall(findPosition,ws.cell(c,3).value)
        print(str)
        if len(str) !=0:
            s= str[0]
            #print(s[0])
            #print(s[1])
            ws.cell(c,6).value=s[0]+s[1]
table.save("./modified.xlsx")