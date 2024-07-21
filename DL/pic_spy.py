import openpyxl
import re
import requests

url = "C:/Users/20783/Downloads/DL/Pic.xlsx"

findPosition = re.compile(r'2(.*?)·(.*?) ')
table = openpyxl.load_workbook(url)
ws = table.active

headers = {
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Cookies": "SINAGLOBAL=2489191138537.923.1715332300235; SCF=An94DzJdfhS1Sy3G4jlQDpAkHRfLv0j4XH9XH-lwdHwrcnoje0RQ-YQlHKPByvjxJWlzLEai4XnRsdISW0jNGpQ.; UOR=www.google.com,open.weibo.com,cn.bing.com; ALF=1719969811; SUB=_2A25LWWtDDeRhGeFH6VIZ9CbFzTyIHXVoF-KLrDV8PUJbkNAbLUrZkW1Ne6m8MXo3kAZzPKCtGSajhOO5l1E6fy-F; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WhZn1_auFGSCIDQsh5ZmqpB5JpX5KMhUgL.FoM4eo5RShn4So52dJLoI7Uuqg_feh5f; XSRF-TOKEN=zzfympsSbM_I5CG4rRIKYfne; _s_tentry=-; Apache=6629412370911.438.1717745952620; ULV=1717745952693:9:3:3:6629412370911.438.1717745952620:1717491279760; PC_TOKEN=d6f82afbe8; WBPSESS=Rq93_CvEn-lUdoAy1RJUUT4YV4CO_ssIKevfbqIfMRFiNsyGQVgV1jOBuJ6D6nQ_ejaIjjENG8Vt0LcqAreUSPS9iEPXQMbxw4V4to9cNbnB12YoTRpiBXR1U3Xxc46yoXMIKGZszYWen366NbvZTA==",
    "Referer": "https://weibo.com/"
}

b = 1

for c in range(1,len(ws['A'])):
    if ws.cell(c,4).value != None:
        response = requests.get(ws.cell(c,4).value,headers= headers)
        content = response.content
        if ws.cell(c,1).value != None:
            userName = str(ws.cell(c,1).value).replace('c\n','')
            b=1
        with open('C:/Users/20783/Downloads/pics/{}.jpg'.format(userName+str(b)),'wb') as f:
            f.write(content)
        b+=1

