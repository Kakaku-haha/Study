from selenium  import webdriver
import openpyxl
from  selenium.webdriver.common.by import By
import time

book = openpyxl.load_workbook("C:/Users/ZH/Downloads/DL/IP.xlsx")
sheet = book.active

driver = webdriver.Chrome()
driver.get("https://passport.weibo.com/sso/signin?entry=miniblog&source=miniblog")  # 打开微博登录界面
time.sleep(10)


for row in range(4000,len(sheet['A'])):
    #try:
    if sheet.cell(row,7).value !=None:
        print("continue")
        continue
    if row % 200 ==0:
        time.sleep(200)

    if sheet.cell(row,6).value !=None:
        driver.get(sheet.cell(row,6).value)
        time.sleep(2)
        ip = driver.find_element(By.XPATH,"//*[@id='app']/div[2]/div[2]/div[2]/main/div/div/div[2]/div[1]/div[1]/div[3]")
        sheet.cell(row,7).value = ip.text
        print(ip.text)
        book.save("./IP2.xlsx")
    #except:print("Noooooh!")
